import os
import sys
import numpy as np
import torch
import torch.nn.functional as F
from torchvision import transforms
sys.path.append('./models/backbone')

import datasets.mvtec as mvtec
from datasets.mvtec import _CLASSNAMES as _CLASSNAMES_mvtec_ad
import datasets.visa as visa
from datasets.visa import _CLASSNAMES as _CLASSNAMES_visa
import datasets.btad as btad
from datasets.btad import _CLASSNAMES as _CLASSNAMES_btad
import datasets.mvtec_jc as mvtec_jc
from datasets.mvtec_jc import _CLASSNAMES as _CLASSNAMES_mvtec_jc
import datasets.reference_jc as re_jc
from datasets.reference_jc import _CLASSNAMES as _CLASSNAMES_re_jc

import models.backbone.open_clip as open_clip
import models.backbone._backbones as _backbones
from models.modules._LNAMD import LNAMD
from models.modules._MSM import MSM
from models.modules._RsCIN import RsCIN
from models.modules._queue import FixedSizeFIFOQueue
from utils.metrics import compute_metrics
from openpyxl import Workbook
from tqdm import tqdm
import pickle
import time
import cv2
import random
import shutil
import warnings
warnings.filterwarnings("ignore")



class MuSc():
    def __init__(self, cfg, seed=0):
        self.cfg = cfg
        self.seed = seed
        self.device = torch.device("cuda:{}".format(cfg['device']) if torch.cuda.is_available() else "cpu")

        self.path = cfg['datasets']['data_path']
        self.dataset = cfg['datasets']['dataset_name']
        self.vis = cfg['testing']['vis']
        self.vis_type = cfg['testing']['vis_type']
        self.save_excel = cfg['testing']['save_excel']
        self.N = cfg['datasets']['BN']
        self.re_path = cfg['datasets']['reference_path']
        self.theshold = cfg['testing']['theshold']
        # the categories to be tested
        self.categories = cfg['datasets']['class_name']
        self.choice = cfg['models']['choice']
        if isinstance(self.categories, str):
            if self.categories.lower() == 'all':
                if self.dataset == 'visa':
                    self.categories = _CLASSNAMES_visa
                elif self.dataset == 'mvtec_ad':
                    self.categories = _CLASSNAMES_mvtec_ad
                elif self.dataset == 'btad':
                    self.categories = _CLASSNAMES_btad
                elif self.dataset == 'mvtec_ad_jc':
                    self.categories = _CLASSNAMES_mvtec_jc
            else:
                self.categories = [self.categories]

        self.model_name = cfg['models']['backbone_name']
        self.image_size = cfg['datasets']['img_resize']
        self.batch_size = cfg['models']['batch_size']
        self.pretrained = cfg['models']['pretrained']
        self.features_list = [l+1 for l in cfg['models']['feature_layers']]
        # self.divide_num = cfg['datasets']['divide_num']
        self.divide_num = 1
        self.r_list = cfg['models']['r_list']
        self.output_dir = os.path.join(cfg['testing']['output_dir'], self.dataset, self.model_name, 'imagesize{}'.format(self.image_size))
        self.refresh = cfg['testing']['refresh']
        os.makedirs(self.output_dir, exist_ok=True)
        self.load_backbone()


    def load_backbone(self):
        if 'dino' in self.model_name:
            # dino or dino_v2
            self.dino_model = _backbones.load(self.model_name)
            self.dino_model.to(self.device)
            self.preprocess = None
        else:
            # clip
            self.clip_model, _, self.preprocess = open_clip.create_model_and_transforms(self.model_name, self.image_size, pretrained=self.pretrained)
            self.clip_model.to(self.device)


    def load_datasets(self, category, divide_num=1, divide_iter = 0, split=mvtec_jc.DatasetSplit.TEST):
        # dataloader
        # re_dataset = re_jc.ReJCDataset(source=self.re_path, split=re_jc.DatasetSplit.TEST,
        #                                     classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess, BN =self.N,
        #                                     random_seed=self.seed)
        if self.dataset == 'visa':
            test_dataset = visa.VisaDataset(source=self.path, split=split,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        elif self.dataset == 'mvtec_ad':
            test_dataset = mvtec.MVTecDataset(source=self.path, split=split,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        elif self.dataset == 'btad':
            test_dataset = btad.BTADDataset(source=self.path, split=split,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)
        elif self.dataset == 'mvtec_jc':
            test_dataset = mvtec_jc.MVTecJCDataset(source=self.path, split=split,
                                            classname=category, resize=self.image_size, imagesize=self.image_size, clip_transformer=self.preprocess,
                                                divide_num=divide_num, divide_iter=divide_iter, random_seed=self.seed)   
        return test_dataset


    def visualization(self, image_path_list, gt_list, pr_px, category):
        def normalization01(img):
            return (img - img.min()) / (img.max() - img.min())
        if self.vis_type == 'single_norm':
            # normalized per image
            for i, path in enumerate(image_path_list):
                anomaly_type = path.split('/')[-2]
                img_name = path.split('/')[-1]
                if anomaly_type not in ['good', 'Normal', 'ok'] and gt_list[i] != 0:
                    save_path = os.path.join(self.output_dir, category, anomaly_type)
                    os.makedirs(save_path, exist_ok=True)
                    save_path = os.path.join(save_path, img_name)
                    anomaly_map = pr_px[i].squeeze()
                    anomaly_map = normalization01(anomaly_map)*255
                    anomaly_map = cv2.applyColorMap(anomaly_map.astype(np.uint8), cv2.COLORMAP_JET)
                    cv2.imwrite(save_path, anomaly_map)
        else:
            # normalized all image
            pr_px = normalization01(pr_px)
            for i, path in enumerate(image_path_list):
                anomaly_type = path.split('\\')[-2]
                img_name = path.split('\\')[-1]
                save_path = os.path.join( self.output_dir, category, anomaly_type)
                os.makedirs(save_path, exist_ok=True)
                #save_path = os.path.join(save_path, img_name)
                
                anomaly_map = pr_px[i].squeeze()
                anomaly_map *= 255
                save_path =  save_path+'/'+str(round(np.max(anomaly_map),2))+'_'+img_name
                anomaly_map = cv2.applyColorMap(anomaly_map.astype(np.uint8), cv2.COLORMAP_JET)
                cv2.imwrite(save_path, anomaly_map)


    def make_category_data(self, category):
        print(category)

        # divide sub-datasets
        divide_num = self.divide_num
        anomaly_maps = torch.tensor([]).double()
        reference_maps = torch.tensor([]).double()
        gt_list = []
        img_masks = []
        class_tokens = []
        image_path_list = []
        scores_cls = []
        start_time_all = time.time()
        dataset_num = 0
        train_dataset  = self.load_datasets(category, divide_num=4, divide_iter=0, split=mvtec_jc.DatasetSplit.TRAIN)
        re_dataloader = torch.utils.data.DataLoader(
                train_dataset,
                batch_size=self.batch_size,
                shuffle=False,
                num_workers=0,
                pin_memory=True,
            )
        


        # 对模板文件夹中的模板图像进行特征提取
        re_patch_tokens_list = []
        for image_all in tqdm(re_dataloader):
            # reference_data = image_all['reference']
            reference_data = image_all["image"]
            with torch.no_grad(), torch.cuda.amp.autocast():
                reference = reference_data.to(torch.float).to(self.device)
                if 'dinov2' in self.model_name:
                    patch_tokens = self.dino_model.get_intermediate_layers(x=reference, n=[l-1 for l in self.features_list], return_class_token=False)
                    image_features = self.dino_model(reference)
                    patch_tokens = [patch_tokens[l].cpu() for l in range(len(self.features_list))]
                    fake_cls = [torch.zeros_like(p)[:, 0:1, :] for p in patch_tokens]
                    patch_tokens = [torch.cat([fake_cls[i], patch_tokens[i]], dim=1) for i in range(len(patch_tokens))]
                else: # clip
                    reference_features, re_patch_tokens = self.clip_model.encode_image(reference, self.features_list)
                    reference_features /= reference_features.norm(dim=-1, keepdim=True)
                    re_patch_tokens = [re_patch_tokens[l].cpu() for l in range(len(self.features_list))]
            re_patch_tokens_list.append(re_patch_tokens)  # (B, L+1, C)
            # re_patch_tokens_list.append(patch_tokens)
            # reference_features = [reference_features[bi].squeeze().cpu().numpy() for bi in range(reference_features.shape[0])]
        print('reference features extracted')

        # LNAMD_re
        re_feature_dim = re_patch_tokens_list[0][0].shape[-1]
        re_anomaly_maps_r = torch.tensor([]).double().to(self.device)
        Z_2 = []
        for r in self.r_list:
            print('aggregation degree: {}'.format(r))
            LNAMD_r = LNAMD(device=self.device, r=r, feature_dim=re_feature_dim, feature_layer=self.features_list)
            re_Z_layers = {}
            for im in range(len(re_patch_tokens_list)):
                re_patch_tokens = [p.to(self.device) for p in re_patch_tokens_list[im]]
                with torch.no_grad(), torch.cuda.amp.autocast():
                    re_features = LNAMD_r._embed(re_patch_tokens)
                    re_features /= re_features.norm(dim=-1, keepdim=True)
                    for l in range(len(self.features_list)):
                        # save the aggregated features
                        if str(l) not in re_Z_layers.keys():
                            re_Z_layers[str(l)] = []
                        re_Z_layers[str(l)].append(re_features[:, :, l, :])
            for l in re_Z_layers.keys():
                Z_2.append(torch.cat(re_Z_layers[l], dim=0).to(self.device))
        R = torch.stack(Z_2, dim=0).to(self.device).permute(1,2,0,3).mean(dim=2)
        R /= R.norm(dim=-1, keepdim=True)
        Reference_queue = FixedSizeFIFOQueue(max_size = self.N + 1)
        for i in range(R.shape[0]):
            Reference_queue.enqueue(R[i].unsqueeze(0))
        # reference_j = [R[i].unsqueeze(0) for i in range(R.shape[0])]
        print('LNAMD_re LNAMD extracted')



        #对每批次图像进行MuSc打分，批次数量：divide_num，批次数：divide_iter 
        divide_iter = 1
        Reference_queue_t = Reference_queue.to_tensor() #这里每次运行可能会拖慢速度
        test_dataset  = self.load_datasets(category, divide_num=divide_num, divide_iter=divide_iter, split=mvtec_jc.DatasetSplit.TEST)
        test_dataloader = torch.utils.data.DataLoader(
            test_dataset,
            batch_size=self.batch_size,
            shuffle=False,
            num_workers=0,
            pin_memory=True,
        )   

        # 对待测图像进行特征提取
        subset_num = len(test_dataset)
        good_queue = FixedSizeFIFOQueue(max_size = self.N + 1)
        good_list = []
        dataset_num += subset_num
        start_time = time.time()
        i = 0
        time_extract_list = []
        time_LMF_list = []
        time_MSM_list = []
        time_MSM_dafen_list = []
        time_MSM_ac_list = []
        time_MSM_chazhi_list = []
        time_MSM_tianjia_list = []
        time_io_list = []
        
        for  image_info in tqdm(test_dataloader):   
        # for image_info in test_dataloader:
            if i > 1:
                end_time_io = time.time()
                time_io_list.append(end_time_io-start_time_io)
            start_time_extract = time.time()

            if isinstance(image_info, dict):
                image = image_info["image"]
                image_path_list.extend(image_info["image_path"])
                img_masks.append(image_info["mask"])
                gt_list.extend(list(image_info["is_anomaly"].numpy()))
            with torch.no_grad(), torch.cuda.amp.autocast():
                input_image = image.to(torch.float).to(self.device)
                if 'dinov2' in self.model_name:
                    patch_tokens = self.dino_model.get_intermediate_layers(x=input_image, n=[l-1 for l in self.features_list], return_class_token=False)
                    image_features = self.dino_model(input_image)
                    patch_tokens = [patch_tokens[l].cpu() for l in range(len(self.features_list))]
                    fake_cls = [torch.zeros_like(p)[:, 0:1, :] for p in patch_tokens]
                    patch_tokens = [torch.cat([fake_cls[i], patch_tokens[i]], dim=1) for i in range(len(patch_tokens))]
                elif 'dino' in self.model_name:
                    patch_tokens_all = self.dino_model.get_intermediate_layers(x=input_image, n=max(self.features_list))
                    image_features = self.dino_model(input_image)
                    patch_tokens = [patch_tokens_all[l-1].cpu() for l in self.features_list]
                else: # clip
                    image_features, patch_tokens = self.clip_model.encode_image(input_image, self.features_list)
                    # image_features /= image_features.norm(dim=-1, keepdim=True)
                    patch_tokens = [patch_tokens[l] for l in range(len(self.features_list))]
            feature_dim = patch_tokens[0].shape[-1]
            anomaly_maps_r = torch.tensor([]).double().to(self.device)
            Z_1 = []

            #测试集特征提取时间
            end_time_extract = time.time()
            extract_time = end_time_extract - start_time_extract
            time_extract_list.append(extract_time)

            start_time_LMF = time.time()
            for r in self.r_list:
                # print('aggregation degree: {}'.format(r))
                LNAMD_r = LNAMD(device=self.device, r=r, feature_dim=feature_dim, feature_layer=self.features_list)
                Z_layers = {}
                patch_tokens = [p for p in patch_tokens]
                with torch.no_grad(), torch.cuda.amp.autocast():
                    features = LNAMD_r._embed(patch_tokens)
                    features /= features.norm(dim=-1, keepdim=True)
                    for l in range(len(self.features_list)):
                        # save the aggregated features
                        if str(l) not in Z_layers.keys():
                            Z_layers[str(l)] = []
                        Z_layers[str(l)].append(features[:, :, l, :])
                for l in Z_layers.keys():
                    Z_1.append(torch.cat(Z_layers[l], dim=0).to(self.device))
            Z = torch.stack(Z_1, dim=0).permute(1,2,0,3).mean(dim=2)
            Z /= Z.norm(dim=-1, keepdim=True)

            #测试集LMF时间
            end_time_LMF = time.time()
            time_LMF = end_time_LMF - start_time_LMF
            time_LMF_list.append(time_LMF)
            
            #MSM互打分
            start_time_MSM = time.time()
            start_time_MSM_dafen = time.time()
            anomaly_maps_msm = MSM(Z=Z, R=Reference_queue_t, device=self.device, topmin_min=0, topmin_max=0.3, choice = self.choice)
            anomaly_maps_r = torch.cat((anomaly_maps_r, anomaly_maps_msm), dim=0).to(self.device)
            del anomaly_maps_msm
            end_time_MSM_dafen = time.time()
            time_MSM_dafen = end_time_MSM_dafen - start_time_MSM_dafen
            time_MSM_dafen_list.append(time_MSM_dafen)
            torch.cuda.empty_cache()

            


            start_time_MSM_ac = time.time()
            B, L = anomaly_maps_r.shape
            top_values, _ = torch.topk(anomaly_maps_r, k=25)
            ac_score_1 = top_values.mean().item()
            # ac_score_1 = torch.max(anomaly_maps_r, dim=1)[0].item()
            scores_cls.append(ac_score_1) #记录图像级异常分数

            end_time_MSM_ac = time.time()
            time_MSM_ac = end_time_MSM_ac - start_time_MSM_ac
            time_MSM_ac_list.append(time_MSM_ac)

            start_time_MSM_chazhi = time.time()
            H = int(np.sqrt(L))
            anomaly_maps_r = F.interpolate(anomaly_maps_r.view(B, 1, H, H),
                                        size=self.image_size, mode='bilinear', align_corners=True)
            anomaly_maps_r = anomaly_maps_r.cpu()

            anomaly_maps = torch.cat((anomaly_maps, anomaly_maps_r), dim=0)
            
            


            torch.cuda.empty_cache()
            a = test_dataset.data_to_iterate[i]
            if ac_score_1 < self.theshold:
                #将检测为正常的图像特征放入参考列表，并从参考列表中取出一个特征替换
                Reference_queue.enqueue(Z)
                Reference_queue_t = Reference_queue.to_tensor()

                #将检测为good的图像保存到good_item文件夹
                target_dir_x = os.path.join(self.output_dir + '/'+ category + '/good_item')
                if not os.path.exists(target_dir_x):
                    os.makedirs(target_dir_x)
                target_x = os.path.join(target_dir_x + '/' + a[1] + os.path.basename(a[2]) )
                good_queue.enqueue(target_x)
                shutil.copy(a[2], target_x)
            
            elif ac_score_1 > self.theshold:
                #将检测为anomaly的图像保存到anomaly_item文件夹
                target_dir_y = os.path.join(self.output_dir + '/'+ category + '/anomaly_item')
                if not os.path.exists(target_dir_y):
                    os.makedirs(target_dir_y)
                target_y = os.path.join(target_dir_y + '/' + a[1] + os.path.basename(a[2]) )
                shutil.copy(a[2], target_y)
            i += 1
            end_time_MSM_chazhi = time.time()
            time_MSM_chazhi = end_time_MSM_chazhi - start_time_MSM_chazhi
            time_MSM_chazhi_list.append(time_MSM_chazhi)

            #热插拔
            if i % self.refresh == 0:
                # 确保目标文件夹存在
                reference_path = os.path.join(self.re_path, category, "train", "reference")
                os.makedirs(reference_path, exist_ok=True)
                # 获取文件列表的长度
                num_files = good_queue.size()
                # 删除目标文件夹中的文件
                for filename in os.listdir(reference_path):
                    file_path = os.path.join(reference_path, filename)
                    if os.path.isfile(file_path):
                        if num_files > 0:  # 仅在有文件要添加时删除
                            os.remove(file_path)  # 删除文件
                            num_files -= 1  # 删除一个文件后，减少计数
                        else:
                            break  # 如果没有文件要添加，停止删除
                # 将文件复制到目标文件夹
                for file_path in good_queue.queue:
                    if os.path.isfile(file_path):  # 确保路径是文件
                        shutil.copy(file_path, reference_path)
                    else:
                        print(f"文件不存在: {file_path}")
                
            if i == 200:
                    print(f"Allocated: {torch.cuda.memory_allocated(self.device) / 1024 ** 2:.3f} MB")
            start_time_io = time.time()
            end_time = time.time()

        print('extract time: {}ms per image'.format((sum(time_extract_list)*1000/subset_num)))
        print('LMF time: {}ms per image'.format((sum(time_LMF_list)*1000/subset_num)))
        print('MSM time: {}ms per image'.format((sum(time_MSM_list)*1000/subset_num)))
        print('MSM_dafen time: {}ms per image'.format((sum(time_MSM_dafen_list)*1000/subset_num)))
        print('MSM_ac time: {}ms per image'.format((sum(time_MSM_ac_list)*1000/subset_num)))
        print('MSM_chazhi time: {}ms per image'.format((sum(time_MSM_chazhi_list)*1000/subset_num)))
        print('io time: {}ms per image'.format((sum(time_io_list))*1000/subset_num))
        print('inference time: {}ms per image'.format((end_time-start_time)*1000/subset_num))



        print('computing metrics...')
        pr_sp = np.array(scores_cls)
        gt_sp = np.array(gt_list)
        gt_px = torch.cat(img_masks, dim=0).numpy().astype(np.int32)
        pr_px = np.array(anomaly_maps)
        image_metric, pixel_metric = compute_metrics(gt_sp, pr_sp, gt_px, pr_px)
        # image_metric = compute_metrics(gt_sp, pr_sp, gt_px, pr_px)
        auroc_sp, f1_sp, ap_sp = image_metric
        auroc_px, f1_px, ap_px, aupro = pixel_metric
        print(category)
        print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp*100, f1_sp*100, ap_sp*100))
        print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px*100, f1_px*100, ap_px*100, aupro*100))

        if self.vis:
            print('visualization...')
            self.visualization(image_path_list, gt_list, pr_px, category)

        return image_metric, pixel_metric
        # return image_metric
    
    
    def save_last_n_files(file_paths, target_folder, n=40):
        # 确保目标文件夹存在
        os.makedirs(target_folder, exist_ok=True)

        # 删除目标文件夹中的所有文件
        for filename in os.listdir(target_folder):
            file_path = os.path.join(target_folder, filename)
            if os.path.isfile(file_path):
                os.remove(file_path)
        
        # 获取倒数 n 个文件路径
        last_n_files = file_paths[-n:]

        # 将文件复制到目标文件夹
        for file_path in last_n_files:
            if os.path.isfile(file_path):  # 确保路径是文件
                shutil.copy(file_path, target_folder)
            else:
                print(f"文件不存在: {file_path}")
    



    def main(self):
        auroc_sp_ls = []
        f1_sp_ls = []
        ap_sp_ls = []
        auroc_px_ls = []
        f1_px_ls = []
        ap_px_ls = []
        aupro_ls = []
        for category in self.categories:
            image_metric, pixel_metric = self.make_category_data(category=category,)
            # image_metric = self.make_category_data(category=category,)
            auroc_sp, f1_sp, ap_sp = image_metric
            auroc_px, f1_px, ap_px, aupro = pixel_metric
            auroc_sp_ls.append(auroc_sp)
            f1_sp_ls.append(f1_sp)
            ap_sp_ls.append(ap_sp)
            auroc_px_ls.append(auroc_px)
            f1_px_ls.append(f1_px)
            ap_px_ls.append(ap_px)
            aupro_ls.append(aupro)
        # mean
        auroc_sp_mean = sum(auroc_sp_ls) / len(auroc_sp_ls)
        f1_sp_mean = sum(f1_sp_ls) / len(f1_sp_ls)
        ap_sp_mean = sum(ap_sp_ls) / len(ap_sp_ls)
        auroc_px_mean = sum(auroc_px_ls) / len(auroc_px_ls)
        f1_px_mean = sum(f1_px_ls) / len(f1_px_ls)
        ap_px_mean = sum(ap_px_ls) / len(ap_px_ls)
        aupro_mean = sum(aupro_ls) / len(aupro_ls)

        for i, category in enumerate(self.categories):
            print(category)
            print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp_ls[i]*100, f1_sp_ls[i]*100, ap_sp_ls[i]*100))
            print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px_ls[i]*100, f1_px_ls[i]*100, ap_px_ls[i]*100, aupro_ls[i]*100))
        print('mean')
        print('image-level, auroc:{}, f1:{}, ap:{}'.format(auroc_sp_mean*100, f1_sp_mean*100, ap_sp_mean*100))
        print('pixel-level, auroc:{}, f1:{}, ap:{}, aupro:{}'.format(auroc_px_mean*100, f1_px_mean*100, ap_px_mean*100, aupro_mean*100))
        
        # save in excel
        if self.save_excel:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "MuSc_results"
            sheet.cell(row=1,column=2,value='auroc_px')
            sheet.cell(row=1,column=3,value='f1_px')
            sheet.cell(row=1,column=4,value='ap_px')
            sheet.cell(row=1,column=5,value='aupro')
            sheet.cell(row=1,column=6,value='auroc_sp')
            sheet.cell(row=1,column=7,value='f1_sp')
            sheet.cell(row=1,column=8,value='ap_sp')
            for col_index in range(2):
                for row_index in range(len(self.categories)):
                    if col_index == 0:
                        sheet.cell(row=row_index+2,column=col_index+1,value=self.categories[row_index])
                    else:
                        sheet.cell(row=row_index+2,column=col_index+1,value=auroc_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+2,value=f1_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+3,value=ap_px_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+4,value=aupro_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+5,value=auroc_sp_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+6,value=f1_sp_ls[row_index]*100)
                        sheet.cell(row=row_index+2,column=col_index+7,value=ap_sp_ls[row_index]*100)
                    if row_index == len(self.categories)-1:
                        if col_index == 0:
                            sheet.cell(row=row_index+3,column=col_index+1,value='mean')
                        else:
                            sheet.cell(row=row_index+3,column=col_index+1,value=auroc_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+2,value=f1_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+3,value=ap_px_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+4,value=aupro_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+5,value=auroc_sp_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+6,value=f1_sp_mean*100)
                            sheet.cell(row=row_index+3,column=col_index+7,value=ap_sp_mean*100)
            workbook.save(os.path.join(self.output_dir, 'results.xlsx'))


